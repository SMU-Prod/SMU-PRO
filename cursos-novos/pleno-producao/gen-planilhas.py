# -*- coding: utf-8 -*-
"""
gen-planilhas.py — gera as 2 planilhas .xlsx (com formulas vivas) do curso PLENO - PRODUTOR / SMU.

FONTE DE VERDADE: _FATOS-VIAGEM.md (levantamento 15/07/2026).
REGRA: nenhum numero de bagagem / franquia / limite pode sair de fora daquele arquivo.
Numero que nao esta la (passagem, hotel, diaria, fretamento) = PREMISSA DO ALUNO, celula amarela.

Nota PT-BR: o .xlsx guarda nome de funcao SEMPRE em ingles (SUM/IF/VLOOKUP/INDEX/MATCH).
O Excel/Sheets em portugues traduz na abertura para SOMA/SE/PROCV/INDICE/CORRESP.
Por isso o script escreve em ingles - e o aluno brasileiro ve em portugues.

Uso: python gen-planilhas.py
"""

import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
SNAP = "snapshot 15/07/2026"

# ---------------------------------------------------------------- estilos
AZUL = "1F3864"
FILL_TITULO = PatternFill("solid", fgColor=AZUL)
FILL_HEADER = PatternFill("solid", fgColor="D6E4F0")
FILL_INPUT = PatternFill("solid", fgColor="FFF9C4")   # amarelo claro = o aluno preenche
FILL_CALC = PatternFill("solid", fgColor="EDEDED")    # cinza = calculado, nao mexer
FILL_OFICIAL = PatternFill("solid", fgColor="E2EFDA")  # verde = dado oficial
FILL_ALERTA = PatternFill("solid", fgColor="FCE4EC")
FILL_TOTAL = PatternFill("solid", fgColor="FFE0B2")

F_TITULO = Font(bold=True, size=14, color="FFFFFF")
F_HEADER = Font(bold=True, size=10, color=AZUL)
F_TOTAL = Font(bold=True, size=11)
F_NOTA = Font(italic=True, size=9, color="666666")
F_SEC = Font(bold=True, size=11, color=AZUL)

THIN = Side(style="thin", color="9E9E9E")
BORDA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MOEDA = 'R$ #,##0.00'
PCT = '0.0%'

AVISO_PRECO = ("Preco de bagagem NAO e tabelado por lei: e contrato acessorio (Res. ANAC 400/2016, art. 13) "
               "e preco livre. Os valores abaixo sao SNAPSHOT de 15/07/2026 das paginas oficiais das cias e "
               "mudam sem aviso. Confira antes de fechar orcamento.")


def titulo(ws, texto, ncols=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=texto)
    c.fill = FILL_TITULO
    c.font = F_TITULO
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26


def nota(ws, row, texto, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.font = F_NOTA
    c.alignment = Alignment(vertical="top", wrap_text=True)


def header(ws, row, labels, col0=1):
    for i, t in enumerate(labels):
        c = ws.cell(row=row, column=col0 + i, value=t)
        c.fill = FILL_HEADER
        c.font = F_HEADER
        c.border = BORDA
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def larguras(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- TABELAS_OFICIAIS
# Todo numero daqui sai de _FATOS-VIAGEM.md. Onde a fonte nao publica, escrevemos
# "nao publicado" - e as formulas tratam isso como alerta, nunca como zero.

FONTE_GOL = "GOL - tabela publicada de bagagem nacional (_FATOS-VIAGEM.md, sec. 2)"
FONTE_AZUL = "Azul - tabela publicada Brasil (_FATOS-VIAGEM.md, sec. 2)"
FONTE_LATAM = "LATAM - preco DINAMICO, nao publicado (_FATOS-VIAGEM.md, sec. 2)"

NP = "nao publicado"

TAB_DESPACHO = [
    # chave, cia, antecedencia, 1a, 2a, 3a-5a, fonte, snapshot
    ("GOL|48h ou mais",   "GOL",   "48h ou mais",   130, 155, 210, FONTE_GOL, SNAP),
    ("GOL|menos de 48h",  "GOL",   "menos de 48h",  165, 200, 280, FONTE_GOL, SNAP),
    ("GOL|no aeroporto",  "GOL",   "no aeroporto",  180, 200, 280, FONTE_GOL, SNAP),
    ("Azul|48h ou mais",  "Azul",  "48h ou mais",   175, 200, 220, FONTE_AZUL, SNAP),
    ("Azul|menos de 48h", "Azul",  "menos de 48h",  200, 220, 280, FONTE_AZUL, SNAP),
    ("Azul|no aeroporto", "Azul",  "no aeroporto",  NP,  NP,  NP,  "Azul nao publica preco de balcao/aeroporto. " + FONTE_AZUL, SNAP),
    ("LATAM|48h ou mais", "LATAM", "48h ou mais",   NP,  NP,  NP,  FONTE_LATAM, SNAP),
    ("LATAM|menos de 48h", "LATAM", "menos de 48h", NP,  NP,  NP,  FONTE_LATAM, SNAP),
    ("LATAM|no aeroporto", "LATAM", "no aeroporto", NP,  NP,  NP,  FONTE_LATAM, SNAP),
]
LIN_DESP_INI = 6
LIN_DESP_FIM = LIN_DESP_INI + len(TAB_DESPACHO) - 1   # 14

TAB_EXCESSO = [
    ("GOL",   350, "Excesso de 32 kg ate 45 kg por volume. Acima de 45 kg NAO transporta.", FONTE_GOL),
    ("Azul",  NP,  "Azul nao publica preco de excesso de peso. Recusa acima de 45 kg (BR) / 32 kg (Europa).", FONTE_AZUL),
    ("LATAM", NP,  "Preco dinamico, nao publicado. Nao transporta acima de 45 kg (32 kg em rotas especificas).", FONTE_LATAM),
]
LIN_EXC_INI = 21
LIN_EXC_FIM = LIN_EXC_INI + len(TAB_EXCESSO) - 1   # 23

TAB_ESPECIAL = [
    ("GOL",   195, "292 cm (A+B+C), ate 23 kg. Mesmo preco em +48h / -48h / aeroporto. Cliente Smiles NAO tem gratuidade; "
                   "tarifa com despacho gratis NAO cobre a diferenciada; a GOL nao se responsabiliza por danos a bagagem diferenciada.", FONTE_GOL),
    ("Azul",  250, "Acima de 158 cm somados ja e ESPECIAL. Preco Brasil.", FONTE_AZUL),
    ("LATAM", NP,  "Limite 300 cm lineares. Preco nao publicado. Violao/guitarra com estojo = bagagem REGULAR (consome a franquia).", FONTE_LATAM),
]
LIN_ESP_INI = 27
LIN_ESP_FIM = LIN_ESP_INI + len(TAB_ESPECIAL) - 1   # 29


def aba_tabelas_oficiais(wb):
    ws = wb.create_sheet("TABELAS_OFICIAIS")
    titulo(ws, "TABELAS OFICIAIS - snapshot 15/07/2026 (nao edite: dado de fonte)", 8)
    nota(ws, 2, AVISO_PRECO, 8)
    ws.row_dimensions[2].height = 28

    ws.cell(row=4, column=1, value="A) DESPACHO NACIONAL - preco por peca, por trecho (R$)").font = F_SEC
    header(ws, 5, ["Chave (cia|antecedencia)", "Companhia", "Antecedencia",
                   "1a peca (R$)", "2a peca (R$)", "3a a 5a peca (R$)", "Fonte", "Snapshot"])
    for i, linha in enumerate(TAB_DESPACHO):
        r = LIN_DESP_INI + i
        for j, v in enumerate(linha):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.border = BORDA
            if j in (3, 4, 5):
                if isinstance(v, (int, float)):
                    c.number_format = MOEDA
                    c.fill = FILL_OFICIAL
                else:
                    c.fill = FILL_ALERTA
                    c.font = Font(italic=True, color="B71C1C")
            if j == 6:
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=8, color="666666")

    ws.cell(row=16, column=1, value="Azul: maximo de 5 pecas de 23 kg por passageiro/trecho. "
                                    "Fonte: _FATOS-VIAGEM.md sec. 2 (Azul Brasil publicado).").font = F_NOTA
    ws.cell(row=17, column=1, value="NAO EXISTE NESTA PLANILHA a faixa de excesso 23->32 kg da GOL: "
                                    "duas paginas oficiais da GOL divergem (R$ 170 x R$ 275). "
                                    "Dado em conflito nao vira celula - vira alerta.").font = Font(italic=True, size=9, color="B71C1C")

    ws.cell(row=LIN_EXC_INI - 2, column=1, value="B) EXCESSO DE PESO - 32 kg a 45 kg, por volume (R$)").font = F_SEC
    header(ws, LIN_EXC_INI - 1, ["Companhia", "Excesso 32-45 kg (R$/volume)", "Observacao", "Fonte"])
    for i, linha in enumerate(TAB_EXCESSO):
        r = LIN_EXC_INI + i
        for j, v in enumerate(linha):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.border = BORDA
            if j == 1:
                if isinstance(v, (int, float)):
                    c.number_format = MOEDA
                    c.fill = FILL_OFICIAL
                else:
                    c.fill = FILL_ALERTA
                    c.font = Font(italic=True, color="B71C1C")
            if j in (2, 3):
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=8, color="666666")

    ws.cell(row=LIN_ESP_INI - 2, column=1, value="C) BAGAGEM ESPECIAL / DIFERENCIADA - case rigido obrigatorio "
                                                 "(bumbo, contrabaixo, tuba, violoncelo) - por volume, por trecho (R$)").font = F_SEC
    header(ws, LIN_ESP_INI - 1, ["Companhia", "Preco (R$/volume)", "Observacao", "Fonte"])
    for i, linha in enumerate(TAB_ESPECIAL):
        r = LIN_ESP_INI + i
        for j, v in enumerate(linha):
            c = ws.cell(row=r, column=1 + j, value=v)
            c.border = BORDA
            if j == 1:
                if isinstance(v, (int, float)):
                    c.number_format = MOEDA
                    c.fill = FILL_OFICIAL
                else:
                    c.fill = FILL_ALERTA
                    c.font = Font(italic=True, color="B71C1C")
            if j in (2, 3):
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=8, color="666666")

    ws.cell(row=LIN_ESP_FIM + 2, column=1, value="D) FRANQUIAS E LIMITES (dado de norma / contrato de cia)").font = F_SEC
    header(ws, LIN_ESP_FIM + 3, ["Tema", "Fato", "Fonte"])
    limites = [
        ("Bagagem de mao - franquia minima", "10 kg. A norma NAO fixa dimensao: dimensao e quantidade sao do CONTRATO DE TRANSPORTE de cada cia.",
         "Res. ANAC 400/2016, art. 14"),
        ("MITO N.1", "55x35x25 cm NAO e 'padrao ANAC'. E contrato de cada companhia.", "Res. ANAC 400/2016, art. 14"),
        ("Bagagem despachada", "Configura CONTRATO ACESSORIO -> franquia gratis NAO e obrigatoria.", "Res. ANAC 400/2016, art. 13"),
        ("Bagagem fora das regras", "Pode ser RECUSADA ou submetida a contrato de transporte de CARGA.", "Res. ANAC 400/2016, art. 15"),
        ("Limite de indenizacao (domestico)", "1.131 DES.", "Res. ANAC 400/2016, art. 17"),
        ("Peso por volume despachado", "23 kg (GOL / Azul / LATAM).", "Cias - " + SNAP),
        ("Dimensao do volume despachado", "LATAM 158 cm lineares | GOL 80x50x28 | Azul 158 cm somados.", "Cias - " + SNAP),
        ("Recusa por peso", "Acima de 45 kg NAO transporta (GOL/Azul/LATAM no Brasil). 32 kg e limite internacional/Europa.", "Cias - " + SNAP),
        ("MITO N.3", "'32 kg e o limite por saude do trabalhador' NAO vale no domestico brasileiro: as tres aceitam ate 45 kg por volume pagando excesso.",
         "Cias - " + SNAP),
        ("Item pessoal", "LATAM 45x35x20 10 kg | GOL 32x43x22 10 kg | Azul 45x35x20 SEM limite de peso.", "Cias - " + SNAP),
        ("Mala de cabine", "LATAM 55x35x25 12 kg (nao inclusa na Basic) | GOL 55x35x25 12 kg (desde 14/10/2025; antes 10 kg) | Azul 55x35x25 (115 cm) 10 kg.",
         "Cias - " + SNAP),
        ("MITO N.2 - PL 5041/2025", "12 kg de mao + 23 kg despachado gratis NAO E LEI. Aprovado na Camara em 28/10/2025, PARADO NO SENADO desde 05/11/2025.",
         "Tramitacao legislativa - " + SNAP),
        ("Azul - objetos magneticos", "Alto-falante, AMPLIFICADOR e CAIXA ACUSTICA sao artigos perigosos: so ate 95 cm somados e 1 por passageiro. "
                                      "Acima disso -> Azul Cargo. (E isto que quebra o plano de levar backline como bagagem.)", "Azul - " + SNAP),
        ("Baterias de litio", "GOL: novas regras a partir de 04/05/2026, so na bagagem de mao, PROIBIDAS na despachada. "
                              "Azul: max. 2 unidades; ate 100 Wh livre; 100-160 Wh com autorizacao previa; acima de 160 Wh proibido; proibido recarregar a bordo.",
         "Cias - " + SNAP),
        ("Instrumento na cabine", "GOL: guitarras e violoes podem ir na cabine como bagagem de mao. "
                                  "Azul: violao/guitarra na cabine SUBSTITUEM a bagagem de mao de 10 kg. "
                                  "LATAM: NAO AFIRMAR - duas paginas oficiais se contradizem.", "Cias - " + SNAP),
        ("Comprar assento para o instrumento", "NAO CONFIRMADO em fonte oficial. Nao usar como plano.", "-"),
        ("RODOVIARIO - franquia minima do bagageiro", "30 kg, 300 dm3, maior dimensao de qualquer item 1 metro. NAO se aplica a microonibus categoria M3.",
         "Res. ANTT 6.033/2023"),
        ("RODOVIARIO - porta-embrulhos (mao)", "5 kg. A norma NAO fixa dimensao em cm. Volume ali e responsabilidade do passageiro e NAO tem indenizacao.",
         "Res. ANTT 6.033/2023"),
        ("RODOVIARIO - excesso de bagagem", "O teto de 0,5%/kg da Res. 1.432/2006 ACABOU. Hoje e SERVICO ACESSORIO: oferta facultativa e PRECO LIVRE - "
                                            "a empresa pode simplesmente RECUSAR. Por isso esta planilha nao tem tabela de excesso rodoviario: e premissa do aluno.",
         "Res. ANTT 6.033/2023"),
        ("ACHADO CRITICO", "A Res. ANTT 1.432/2006 esta REVOGADA pela 6.033/2023 - e ate a pagina da propria ANTT no ar continua desatualizada citando a norma revogada. Nao cite a 1.432.",
         "Res. ANTT 6.033/2023"),
        ("RODOVIARIO - equipamento fora do padrao", "Avisar a empresa com no minimo 24 h de antecedencia (dever do usuario) - e ainda assim ela NAO e obrigada a aceitar.",
         "Res. ANTT 6.033/2023"),
        ("FRETAMENTO - modalidade da equipe", "EVENTUAL (congressos, competicoes). Circuito fechado: grupo com motivacao comum sai da origem, percorre e volta a origem no mesmo veiculo. "
                                              "VENDA DE BILHETE INDIVIDUAL E PROIBIDA.", "Res. ANTT 4.777/2015"),
        ("FRETAMENTO - documentos a bordo", "Licenca de Viagem + relacao de passageiros (art. 23). Tambem CRLV, CSV e apolice do seguro RC (art. 31 §3). "
                                            "'CITV' NAO EXISTE em fonte oficial.", "Res. ANTT 4.777/2015"),
        ("FRETAMENTO - troca de equipe", "Permitida substituicao de ATE 20% dos passageiros antes da partida.", "Res. ANTT 4.777/2015, art. 36"),
    ]
    r = LIN_ESP_FIM + 4
    for tema, fato, fonte in limites:
        ws.cell(row=r, column=1, value=tema).font = Font(bold=True, size=9)
        ws.cell(row=r, column=1).border = BORDA
        c = ws.cell(row=r, column=2, value=fato)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=9)
        c.border = BORDA
        f = ws.cell(row=r, column=3, value=fonte)
        f.font = Font(size=8, color="666666")
        f.border = BORDA
        ws.row_dimensions[r].height = 30
        r += 1

    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)
    c = ws.cell(row=r + 1, column=1, value="TODO numero desta aba veio de _FATOS-VIAGEM.md (levantamento 15/07/2026). "
                                           "Onde a fonte oficial nao publica ou se contradiz, esta escrito 'nao publicado' de proposito - "
                                           "e as formulas das outras abas devolvem alerta em vez de um numero inventado.")
    c.font = Font(italic=True, size=9, color="B71C1C")
    c.alignment = Alignment(wrap_text=True)

    larguras(ws, {"A": 34, "B": 46, "C": 30, "D": 16, "E": 16, "F": 16, "G": 34, "H": 16})
    ws.freeze_panes = "A6"
    return ws


# chaves de intervalo usadas nas formulas (mesmas nas duas planilhas)
RNG_CHAVE = f"TABELAS_OFICIAIS!$A${LIN_DESP_INI}:$A${LIN_DESP_FIM}"
RNG_PRECOS = f"TABELAS_OFICIAIS!$D${LIN_DESP_INI}:$F${LIN_DESP_FIM}"
RNG_EXC_CIA = f"TABELAS_OFICIAIS!$A${LIN_EXC_INI}:$A${LIN_EXC_FIM}"
RNG_EXC_VAL = f"TABELAS_OFICIAIS!$B${LIN_EXC_INI}:$B${LIN_EXC_FIM}"
RNG_ESP_CIA = f"TABELAS_OFICIAIS!$A${LIN_ESP_INI}:$A${LIN_ESP_FIM}"
RNG_ESP_VAL = f"TABELAS_OFICIAIS!$B${LIN_ESP_INI}:$B${LIN_ESP_FIM}"


def preco_peca(chave_cell, coluna):
    """INDEX+MATCH do preco da n-esima peca. Devolve texto 'nao publicado' se a cia nao publica."""
    return f'=IFERROR(INDEX({RNG_PRECOS},MATCH({chave_cell},{RNG_CHAVE},0),{coluna}),"chave invalida")'


def preco_excesso(cia_cell):
    return f'=IFERROR(INDEX({RNG_EXC_VAL},MATCH({cia_cell},{RNG_EXC_CIA},0)),"chave invalida")'


def preco_especial(cia_cell):
    return f'=IFERROR(INDEX({RNG_ESP_VAL},MATCH({cia_cell},{RNG_ESP_CIA},0)),"chave invalida")'


# ================================================================
# PLANILHA 1 - COMPARADOR AVIAO x ONIBUS
# ================================================================
def planilha_comparador():
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- PREMISSAS ----------------
    ws = wb.create_sheet("PREMISSAS")
    titulo(ws, "PREMISSAS - preencha SO as celulas AMARELAS. Todas as outras abas leem daqui.", 5)
    nota(ws, 2, "Regra da casa: uma celula, uma verdade. O numero de pessoas mora em B4 e SO em B4 - "
                "AEREO, RODOVIARIO, FRETAMENTO e RESUMO apontam para ela com referencia absoluta ($B$4). "
                "Mudou a comitiva? Muda B4 e a planilha inteira se recalcula. "
                "AMARELO = voce preenche | VERDE = dado de fonte oficial | CINZA = formula, nao digite por cima.", 5)
    ws.row_dimensions[2].height = 46
    header(ws, 4, ["Premissa", "Valor", "Unidade", "Origem do dado", "Por que isso importa"])

    prem = [
        # row, label, valor, unidade, origem, porque, fmt, tipo
        (5, "Nº de pessoas na comitiva", 8, "pessoas", "PREMISSA DO ALUNO", "Multiplica quase tudo. Erre aqui e erre o orçamento inteiro.", None, "in"),
        (6, "Nº de trechos (ida e volta = 2)", 2, "trechos", "PREMISSA DO ALUNO", "Bagagem se paga POR TRECHO, não por viagem. É o esquecimento clássico.", '0', "in"),
        (7, "Volumes despachados por pessoa (por trecho)", 1, "volumes", "PREMISSA DO ALUNO", "A 2ª e a 3ª peça custam mais caro que a 1ª.", '0', "in"),
        (8, "Peso de cada volume despachado", 23, "kg",
            "PREMISSA DO ALUNO — vem preenchida com 23 kg, que é o DADO OFICIAL da franquia por volume (GOL/Azul/LATAM, snapshot 15/07/2026)",
            "Acima de 32 kg entra excesso. Acima de 45 kg a cia RECUSA e vira carga (Res. ANAC 400/2016, art. 15).", '0', "in"),
        (9, "Companhia aérea", "GOL", "lista", "ESCOLHA (lista suspensa)", "LATAM não publica preço de despacho: é dinâmico. Escolha LATAM e veja a planilha avisar em vez de inventar.", None, "in"),
        (10, "Antecedência da compra do despacho", "48h ou mais", "lista", "ESCOLHA (lista suspensa)", "Comprar no aeroporto custa mais caro. Na GOL: R$ 130 antecipado x R$ 180 no balcão.", None, "in"),
        (11, "Preço médio da passagem aérea (ida+volta, por pessoa)", 1200, "R$", "PREMISSA DO ALUNO — coloque a SUA cotação",
             "Não existe tabela de passagem: preço é livre e muda por hora. O 1.200 aqui é só exemplo.", MOEDA, "in"),
        (12, "Preço da passagem rodoviária (ida+volta, por pessoa)", 380, "R$", "PREMISSA DO ALUNO — coloque a SUA cotação",
             "Idem. Exemplo — substitua.", MOEDA, "in"),
        (13, "Custo do fretamento (veículo fechado, ida+volta)", 9800, "R$", "PREMISSA DO ALUNO — coloque a SUA cotação",
             "Custo do CONTRATO, não por pessoa. Exemplo — substitua.", MOEDA, "in"),
        (14, "Distância (só de ida)", 1000, "km", "PREMISSA DO ALUNO", "Serve para você conferir se as horas de estrada fazem sentido.", '#,##0', "in"),
        (15, "Horas de viagem porta a porta — AÉREO", 6, "h", "PREMISSA DO ALUNO", "Inclui chegar 2 h antes, voo, esteira e traslado.", '0.0', "in"),
        (16, "Horas de viagem porta a porta — RODOVIÁRIO", 16, "h", "PREMISSA DO ALUNO", "É aqui que nasce a diária extra que ninguém orça.", '0.0', "in"),
        (17, "Diária da equipe", 350, "R$/pessoa/dia", "PREMISSA DO ALUNO", "O dia a mais na estrada é dia de cachê. Não é de graça.", MOEDA, "in"),
        (18, "Diária de hotel", 220, "R$/pessoa/noite", "PREMISSA DO ALUNO", "A noite extra que a estrada gera.", MOEDA, "in"),
        (19, "Custo de traslado por trecho", 60, "R$/pessoa/trecho", "PREMISSA DO ALUNO", "Van/carro casa-aeroporto e aeroporto-hotel.", MOEDA, "in"),
        (20, "Nº de traslados — AÉREO", 4, "traslados/pessoa", "PREMISSA DO ALUNO", "Casa→aeroporto, aeroporto→hotel, e a volta.", '0', "in"),
        (21, "Nº de traslados — RODOVIÁRIO", 2, "traslados/pessoa", "PREMISSA DO ALUNO", "Rodoviária→hotel, ida e volta.", '0', "in"),
        (22, "Nº de traslados — FRETAMENTO", 0, "traslados/pessoa", "PREMISSA DO ALUNO", "O fretado costuma ser porta a porta. Se não for, ajuste.", '0', "in"),
        (23, "Excesso de bagagem RODOVIÁRIO (verba total estimada)", 0, "R$", "PREMISSA DO ALUNO — NÃO EXISTE TABELA",
             "Res. ANTT 6.033/2023: excesso virou serviço acessório, oferta facultativa e PREÇO LIVRE — a empresa pode simplesmente recusar. "
             "Não há número oficial para cravar aqui. Cote com a empresa.", MOEDA, "in"),
        (24, "Orçamento aprovado para transporte", 20000, "R$", "PREMISSA DO ALUNO — é o que o cliente aprovou", "É contra isto que o RESUMO mede a folga.", MOEDA, "in"),
    ]
    for row, label, valor, unid, origem, porque, fmt, tipo in prem:
        ws.cell(row=row, column=1, value=label).border = BORDA
        c = ws.cell(row=row, column=2, value=valor)
        c.fill = FILL_INPUT if tipo == "in" else FILL_OFICIAL
        c.border = BORDA
        c.font = Font(bold=True)
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=origem)
        o.border = BORDA
        o.alignment = Alignment(wrap_text=True, vertical="top")
        o.font = Font(size=9, color="B71C1C" if tipo == "in" else "1B5E20")
        p = ws.cell(row=row, column=5, value=porque)
        p.border = BORDA
        p.alignment = Alignment(wrap_text=True, vertical="top")
        p.font = Font(size=9, color="666666")
        ws.row_dimensions[row].height = 32
        c.comment = Comment(f"{origem}\n\n{porque}", "SMU - Pleno Produtor")

    # bloco calculado
    ws.cell(row=26, column=1, value="CALCULADO A PARTIR DAS PREMISSAS (não digite por cima)").font = F_SEC
    calc = [
        (27, "Chave de consulta da tabela oficial", '=$B$9&"|"&$B$10', "texto",
         "FÓRMULA — monta a chave que o ÍNDICE+CORRESP usa em TABELAS_OFICIAIS", None),
        (28, "Preço da 1ª peça despachada (resolvido)", preco_peca("$B$27", 1), "R$/peça/trecho",
         "TABELAS_OFICIAIS — dado oficial, snapshot 15/07/2026", MOEDA),
        (29, "Preço da 2ª peça despachada (resolvido)", preco_peca("$B$27", 2), "R$/peça/trecho",
         "TABELAS_OFICIAIS — dado oficial, snapshot 15/07/2026", MOEDA),
        (30, "Preço da 3ª a 5ª peça (resolvido)", preco_peca("$B$27", 3), "R$/peça/trecho",
         "TABELAS_OFICIAIS — dado oficial, snapshot 15/07/2026", MOEDA),
        (31, "Preço do excesso 32–45 kg (resolvido)", preco_excesso("$B$9"), "R$/volume/trecho",
         "TABELAS_OFICIAIS — só a GOL publica (R$ 350)", MOEDA),
        (32, "Dias extras de equipe que a estrada gera", "=ROUNDUP(($B$16-$B$15)*2/24,0)", "dias",
         "FÓRMULA — modelo: cada 24 h a mais de estrada (ida+volta) = 1 dia extra de equipe", '0'),
        (33, "Noites extras de hotel que a estrada gera", "=$B$32", "noites",
         "FÓRMULA — aponta para B32: uma célula, uma verdade", '0'),
        (34, "Passageiros que podem ser substituídos no fretamento (20%)", "=ROUNDDOWN($B$5*0.2,0)", "pessoas",
         "Res. ANTT 4.777/2015, art. 36 — substituição de até 20% antes da partida", '0'),
    ]
    for row, label, f, unid, origem, fmt in calc:
        ws.cell(row=row, column=1, value=label).border = BORDA
        c = ws.cell(row=row, column=2, value=f)
        c.fill = FILL_CALC
        c.border = BORDA
        c.font = Font(bold=True)
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=origem)
        o.border = BORDA
        o.alignment = Alignment(wrap_text=True, vertical="top")
        o.font = Font(size=9, color="1B5E20")
        ws.row_dimensions[row].height = 30
        c.comment = Comment(origem, "SMU - Pleno Produtor")

    dv_cia = DataValidation(type="list", formula1='"GOL,Azul,LATAM"', allow_blank=False)
    dv_cia.error = "Escolha GOL, Azul ou LATAM."
    dv_cia.errorTitle = "Companhia inválida"
    ws.add_data_validation(dv_cia)
    dv_cia.add(ws["B9"])

    dv_ant = DataValidation(type="list", formula1='"48h ou mais,menos de 48h,no aeroporto"', allow_blank=False)
    dv_ant.error = "Escolha: 48h ou mais / menos de 48h / no aeroporto."
    dv_ant.errorTitle = "Antecedência inválida"
    ws.add_data_validation(dv_ant)
    dv_ant.add(ws["B10"])

    larguras(ws, {"A": 48, "B": 16, "C": 20, "D": 42, "E": 44})
    ws.freeze_panes = "A5"

    # ---------------- funcao auxiliar de aba de custo ----------------
    def bloco_custo(ws, linhas, r_ini=5, r_fim=16):
        header(ws, 4, ["Item de custo", "Qtde", "Unidade", "Valor unitário (R$)", "Total (R$)", "Fonte / observação"])
        for (r, item, qtde, unid, vunit, total, obs) in linhas:
            ws.cell(row=r, column=1, value=item).border = BORDA
            c = ws.cell(row=r, column=2, value=qtde); c.border = BORDA; c.number_format = '#,##0.##'
            ws.cell(row=r, column=3, value=unid).border = BORDA
            c = ws.cell(row=r, column=4, value=vunit); c.border = BORDA; c.number_format = MOEDA
            c = ws.cell(row=r, column=5, value=total); c.border = BORDA; c.number_format = MOEDA
            o = ws.cell(row=r, column=6, value=obs); o.border = BORDA
            o.alignment = Alignment(wrap_text=True, vertical="top"); o.font = Font(size=9, color="666666")
            ws.row_dimensions[r].height = 30
        # folga
        ultimo = max(l[0] for l in linhas)
        for r in range(ultimo + 1, r_fim + 1):
            for col in range(1, 7):
                ws.cell(row=r, column=col).border = BORDA
        c = ws.cell(row=ultimo + 1, column=1, value="↓ linhas livres — acrescente itens aqui")
        c.font = Font(italic=True, size=9, color="1B5E20")

    # ---------------- AEREO ----------------
    ws = wb.create_sheet("AEREO")
    titulo(ws, "CUSTO AÉREO — comitiva", 6)
    nota(ws, 2, "Toda célula calculada é FÓRMULA. Nada aqui é digitado à mão. Os preços de bagagem vêm de "
                "TABELAS_OFICIAIS por ÍNDICE+CORRESP; o resto vem de PREMISSAS por referência absoluta ($B$5).", 6)
    P = "PREMISSAS!"
    bloco_custo(ws, [
        (5, "Passagem aérea (ida+volta)", f"={P}$B$5", "pessoas", f"={P}$B$11", "=IFERROR(B5*D5,0)",
         "PREMISSA DO ALUNO. Não existe tabela de passagem — preço é livre."),
        (6, "Despacho — 1ª peça", f'=IF({P}$B$7>=1,{P}$B$5*{P}$B$6,0)', "peças × trechos", f"={P}$B$28",
         '=IF(B6=0,0,IF(ISNUMBER(D6),B6*D6,"n/d — cotar"))',
         "TABELAS_OFICIAIS (snapshot 15/07/2026). Res. ANAC 400/2016 art. 13: despacho é contrato acessório — não há franquia grátis obrigatória."),
        (7, "Despacho — 2ª peça", f'=IF({P}$B$7>=2,{P}$B$5*{P}$B$6,0)', "peças × trechos", f"={P}$B$29",
         '=IF(B7=0,0,IF(ISNUMBER(D7),B7*D7,"n/d — cotar"))', "TABELAS_OFICIAIS. A 2ª peça é mais cara que a 1ª."),
        (8, "Despacho — 3ª a 5ª peça", f'=MAX(0,MIN({P}$B$7,5)-2)*{P}$B$5*{P}$B$6', "peças × trechos", f"={P}$B$30",
         '=IF(B8=0,0,IF(ISNUMBER(D8),B8*D8,"n/d — cotar"))', "TABELAS_OFICIAIS. Azul: máximo 5 peças de 23 kg por passageiro/trecho."),
        (9, "Excesso de peso (32–45 kg)", f'=IF(AND({P}$B$8>32,{P}$B$8<=45),{P}$B$7*{P}$B$5*{P}$B$6,0)', "volumes × trechos", f"={P}$B$31",
         '=IF(B9=0,0,IF(ISNUMBER(D9),B9*D9,"n/d — cotar"))',
         "TABELAS_OFICIAIS. Só a GOL publica (R$ 350). A faixa 23→32 kg da GOL NÃO está aqui: duas páginas oficiais divergem (R$ 170 × R$ 275)."),
        (10, "Traslado", f"={P}$B$5*{P}$B$20", "pessoas × traslados", f"={P}$B$19", "=IFERROR(B10*D10,0)",
         "PREMISSA DO ALUNO."),
    ])
    c = ws.cell(row=18, column=1, value="TOTAL AÉREO"); c.font = F_TOTAL; c.fill = FILL_TOTAL; c.border = BORDA
    c = ws.cell(row=18, column=5, value="=SUM(E5:E16)")
    c.font = F_TOTAL; c.fill = FILL_TOTAL; c.border = BORDA; c.number_format = MOEDA
    c.comment = Comment("A SOMA vai até a linha 16 de propósito, com folga: você acrescenta item nas linhas livres "
                        "e o total acompanha. O erro nº 1 de planilha de produtora é a SOMA que não alcança a última linha.",
                        "SMU - Pleno Produtor")
    ws.cell(row=18, column=6, value="SOMA(E5:E16) — com folga de linhas de propósito.").font = F_NOTA
    ws.cell(row=19, column=1, value="Custo aéreo por pessoa").font = Font(bold=True)
    c = ws.cell(row=19, column=5, value=f"=IFERROR($E$18/{P}$B$5,0)"); c.number_format = MOEDA; c.font = Font(bold=True)
    ws.cell(row=20, column=1, value="Bagagem despachada — subtotal").font = Font(bold=True)
    c = ws.cell(row=20, column=5, value="=SUM(E6:E9)"); c.number_format = MOEDA; c.font = Font(bold=True)
    ws.cell(row=20, column=6, value="É este número que o fretamento faz desaparecer.").font = F_NOTA

    ws.cell(row=22, column=1, value="ALERTAS AUTOMÁTICOS").font = F_SEC
    alertas = [
        (23, '=IF(COUNTA(E5:E16)>COUNT(E5:E16),"⚠ Há item de bagagem SEM preço oficial publicado (a cia escolhida não publica). '
             'A SOMA está IGNORANDO esse item em silêncio — cote com a companhia antes de fechar.","ok — todos os itens têm preço oficial")'),
        (24, f'=IF(AND({P}$B$8>23,{P}$B$8<=32),"⚠ Faixa 23–32 kg: duas páginas oficiais da GOL divergem (R$ 170 × R$ 275). '
             f'Dado em conflito não vira número — cote com a cia.","")'),
        (25, f'=IF({P}$B$8>45,"⛔ RECUSADO: acima de 45 kg a cia não transporta como bagagem — vira contrato de CARGA (Res. ANAC 400/2016, art. 15).","")'),
        (26, f'=IF({P}$B$9="LATAM","⚠ LATAM não publica preço de despacho (é dinâmico, varia por rota e antecedência). '
             f'A planilha se recusa a inventar — cote.","")'),
        (27, f'=IF({P}$B$6<2,"⚠ Você está pagando bagagem em 1 trecho só. Despacho se paga POR TRECHO: ida e volta = 2.","")'),
    ]
    for r, f in alertas:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=f)
        c.fill = FILL_ALERTA
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.font = Font(size=9, bold=True, color="B71C1C")
        ws.row_dimensions[r].height = 26
    larguras(ws, {"A": 34, "B": 12, "C": 18, "D": 18, "E": 16, "F": 58})
    ws.freeze_panes = "A5"

    # ---------------- RODOVIARIO ----------------
    ws = wb.create_sheet("RODOVIARIO")
    titulo(ws, "CUSTO RODOVIÁRIO — linha regular", 6)
    nota(ws, 2, "A lição desta aba está nas linhas 7 e 8. O ônibus parece barato porque quase todo orçamento "
                "para na passagem. A viagem longa gera dia de equipe e noite de hotel — e isso é dinheiro seu.", 6)
    bloco_custo(ws, [
        (5, "Passagem rodoviária (ida+volta)", f"={P}$B$5", "pessoas", f"={P}$B$12", "=IFERROR(B5*D5,0)",
         "PREMISSA DO ALUNO."),
        (6, "Excesso de bagagem (verba)", 1, "verba", f"={P}$B$23", "=IFERROR(B6*D6,0)",
         "SEM TABELA. Res. ANTT 6.033/2023: excesso virou serviço acessório — oferta facultativa e PREÇO LIVRE, "
         "e a empresa pode simplesmente RECUSAR. Franquia mínima: 30 kg / 300 dm³ / maior dimensão 1 m."),
        (7, "Diária extra da equipe (gerada pela estrada)", f"={P}$B$5*{P}$B$32", "pessoas × dias", f"={P}$B$17", "=IFERROR(B7*D7,0)",
         "O dia a mais na estrada é dia de cachê. Esta é a linha que some do orçamento de ônibus."),
        (8, "Hotel extra (pernoite gerado pela estrada)", f"={P}$B$5*{P}$B$33", "pessoas × noites", f"={P}$B$18", "=IFERROR(B8*D8,0)",
         "Idem. Se a estrada empurra a chegada para outro dia, alguém dorme em algum lugar."),
        (9, "Traslado", f"={P}$B$5*{P}$B$21", "pessoas × traslados", f"={P}$B$19", "=IFERROR(B9*D9,0)",
         "PREMISSA DO ALUNO."),
    ])
    c = ws.cell(row=18, column=1, value="TOTAL RODOVIÁRIO"); c.font = F_TOTAL; c.fill = FILL_TOTAL; c.border = BORDA
    c = ws.cell(row=18, column=5, value="=SUM(E5:E16)")
    c.font = F_TOTAL; c.fill = FILL_TOTAL; c.border = BORDA; c.number_format = MOEDA
    ws.cell(row=18, column=6, value="SOMA(E5:E16) — com folga de linhas de propósito.").font = F_NOTA
    ws.cell(row=19, column=1, value="Custo rodoviário por pessoa").font = Font(bold=True)
    c = ws.cell(row=19, column=5, value=f"=IFERROR($E$18/{P}$B$5,0)"); c.number_format = MOEDA; c.font = Font(bold=True)

    ws.cell(row=21, column=1, value="A ILUSÃO DO ÔNIBUS BARATO").font = F_SEC
    ws.cell(row=22, column=1, value="Total se você ESQUECER a diária e o hotel (o orçamento errado)")
    c = ws.cell(row=22, column=5, value="=$E$18-$E$7-$E$8"); c.number_format = MOEDA
    ws.cell(row=23, column=1, value="Quanto você teria escondido de si mesmo")
    c = ws.cell(row=23, column=5, value="=$E$7+$E$8"); c.number_format = MOEDA; c.font = Font(bold=True, color="B71C1C")
    ws.cell(row=24, column=1, value="…em % do total rodoviário")
    c = ws.cell(row=24, column=5, value="=IFERROR(($E$7+$E$8)/$E$18,0)"); c.number_format = PCT; c.font = Font(bold=True, color="B71C1C")
    ws.merge_cells(start_row=25, start_column=1, end_row=25, end_column=6)
    c = ws.cell(row=25, column=1, value='=IF($E$18-$E$7-$E$8<AEREO!$E$18,"⚠ Sem a diária e o hotel, o ônibus PARECE mais barato que o avião. '
                                        'Com eles, olhe o RESUMO antes de decidir.","")')
    c.fill = FILL_ALERTA; c.font = Font(size=9, bold=True, color="B71C1C"); c.alignment = Alignment(wrap_text=True)
    larguras(ws, {"A": 40, "B": 12, "C": 18, "D": 18, "E": 16, "F": 58})
    ws.freeze_panes = "A5"

    # ---------------- FRETAMENTO ----------------
    ws = wb.create_sheet("FRETAMENTO")
    titulo(ws, "CUSTO FRETAMENTO — eventual, circuito fechado (Res. ANTT 4.777/2015)", 6)
    nota(ws, 2, "Fretamento EVENTUAL é a modalidade de equipe de evento (congressos, competições). "
                "Circuito fechado: o grupo sai da origem, percorre e volta à origem no mesmo veículo. "
                "Custo é do CONTRATO, não por pessoa — e não há venda de bilhete individual (é proibida).", 6)
    bloco_custo(ws, [
        (5, "Fretamento eventual (veículo fechado, ida+volta)", 1, "contrato", f"={P}$B$13", "=IFERROR(B5*D5,0)",
         "PREMISSA DO ALUNO — coloque a sua cotação. Exija Termo de Autorização da ANTT (consulte o CNPJ no site), "
         "Licença de Viagem e relação de passageiros a bordo (art. 23)."),
        (6, "Bilhete individual", 0, "bilhetes", 0, "=IFERROR(B6*D6,0)",
         "Res. ANTT 4.777/2015: a venda de bilhete individual no fretamento é PROIBIDA. Fica zero de propósito."),
        (7, "Diária extra da equipe (gerada pela estrada)", f"={P}$B$5*{P}$B$32", "pessoas × dias", f"={P}$B$17", "=IFERROR(B7*D7,0)",
         "A estrada é a mesma do ônibus de linha: o dia extra continua existindo."),
        (8, "Hotel extra (pernoite gerado pela estrada)", f"={P}$B$5*{P}$B$33", "pessoas × noites", f"={P}$B$18", "=IFERROR(B8*D8,0)",
         "Idem."),
        (9, "Traslado", f"={P}$B$5*{P}$B$22", "pessoas × traslados", f"={P}$B$19", "=IFERROR(B9*D9,0)",
         "Premissa 0 por padrão: o fretado costuma ser porta a porta. Se não for, ajuste em PREMISSAS!B22."),
    ])
    c = ws.cell(row=18, column=1, value="TOTAL FRETAMENTO"); c.font = F_TOTAL; c.fill = FILL_TOTAL; c.border = BORDA
    c = ws.cell(row=18, column=5, value="=SUM(E5:E16)")
    c.font = F_TOTAL; c.fill = FILL_TOTAL; c.border = BORDA; c.number_format = MOEDA
    ws.cell(row=18, column=6, value="SOMA(E5:E16) — com folga de linhas de propósito.").font = F_NOTA
    ws.cell(row=19, column=1, value="Custo do fretamento por pessoa").font = Font(bold=True)
    c = ws.cell(row=19, column=5, value=f"=IFERROR($E$18/{P}$B$5,0)"); c.number_format = MOEDA; c.font = Font(bold=True)

    ws.cell(row=21, column=1, value="O QUE O FRETAMENTO DISPENSA (informativo — não entra no total)").font = F_SEC
    ws.cell(row=22, column=1, value="Bagagem despachada que você deixa de pagar (= subtotal de despacho do AÉREO)")
    c = ws.cell(row=22, column=5, value="=AEREO!$E$20"); c.number_format = MOEDA; c.font = Font(bold=True, color="1B5E20")
    ws.cell(row=22, column=6, value="No fretado o bagageiro é do veículo que você contratou — não há tabela de peça despachada "
                                    "por passageiro nem cobrança por trecho.").font = F_NOTA
    ws.cell(row=23, column=1, value="Traslados que você deixa de pagar em relação ao aéreo")
    c = ws.cell(row=23, column=5, value=f"=IFERROR(({P}$B$20-{P}$B$22)*{P}$B$5*{P}$B$19,0)"); c.number_format = MOEDA; c.font = Font(bold=True, color="1B5E20")
    ws.cell(row=24, column=1, value="Passageiros que você pode trocar antes da partida (até 20%)")
    c = ws.cell(row=24, column=5, value=f"={P}$B$34"); c.number_format = '0'; c.font = Font(bold=True, color="1B5E20")
    ws.cell(row=24, column=6, value="Res. ANTT 4.777/2015, art. 36 — margem real para troca de equipe.").font = F_NOTA
    ws.merge_cells(start_row=26, start_column=1, end_row=26, end_column=6)
    c = ws.cell(row=26, column=1, value="Antes de assinar: consulte a autorização da empresa por CNPJ no site da ANTT, confira o selo no ônibus, "
                                        "exija a Licença de Viagem e inspecione veículo e motorista. Documentos a bordo (art. 23): Licença de Viagem "
                                        "+ relação de passageiros; também CRLV, CSV e apólice do seguro RC (art. 31 §3º). 'CITV' não existe — se alguém "
                                        "te mostrar um, é papel inventado.")
    c.font = Font(size=9, color="1B5E20"); c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[26].height = 44
    larguras(ws, {"A": 46, "B": 12, "C": 18, "D": 18, "E": 16, "F": 58})
    ws.freeze_panes = "A5"

    # ---------------- RESUMO ----------------
    ws = wb.create_sheet("RESUMO")
    titulo(ws, "RESUMO — qual modal, e cabe no orçamento?", 5)
    nota(ws, 2, "Nada aqui é digitado. Tudo aponta para AEREO, RODOVIARIO, FRETAMENTO e PREMISSAS.", 5)
    header(ws, 4, ["Indicador", "AÉREO", "RODOVIÁRIO", "FRETAMENTO", "Observação"])

    linhas = [
        (5, "Custo total (R$)", "=AEREO!$E$18", "=RODOVIARIO!$E$18", "=FRETAMENTO!$E$18", MOEDA, "Cada um vem da SOMA da sua aba."),
        (6, "Custo por pessoa (R$)", f"=IFERROR(B5/{P}$B$5,0)", f"=IFERROR(C5/{P}$B$5,0)", f"=IFERROR(D5/{P}$B$5,0)", MOEDA, ""),
        (7, "Horas porta a porta", f"={P}$B$15", f"={P}$B$16", f"={P}$B$16", '0.0', "Estrada é estrada: linha e fretado gastam o mesmo tempo."),
        (8, "Horas-equipe consumidas (ida+volta)", f"=B7*{P}$B$5*{P}$B$6", f"=C7*{P}$B$5*{P}$B$6", f"=D7*{P}$B$5*{P}$B$6", '#,##0', "Tempo de gente é custo, mesmo quando não vira linha."),
        (9, "Cabe no orçamento?", f'=IF(B5<={P}$B$24,"CABE","ESTOURA")', f'=IF(C5<={P}$B$24,"CABE","ESTOURA")', f'=IF(D5<={P}$B$24,"CABE","ESTOURA")', None, "Contra PREMISSAS!B24."),
        (10, "Folga contra o orçamento (R$)", f"={P}$B$24-B5", f"={P}$B$24-C5", f"={P}$B$24-D5", MOEDA, "Negativo = estourou."),
        (11, "Folga contra o orçamento (%)", f"=IFERROR(({P}$B$24-B5)/{P}$B$24,0)", f"=IFERROR(({P}$B$24-C5)/{P}$B$24,0)", f"=IFERROR(({P}$B$24-D5)/{P}$B$24,0)", PCT, ""),
    ]
    for row, item, a, b, c_, fmt, obs in linhas:
        ws.cell(row=row, column=1, value=item).border = BORDA
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i, f in enumerate([a, b, c_]):
            cc = ws.cell(row=row, column=2 + i, value=f)
            cc.border = BORDA
            if fmt:
                cc.number_format = fmt
        o = ws.cell(row=row, column=5, value=obs); o.border = BORDA
        o.font = Font(size=9, color="666666"); o.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 26
    for row in (5, 6):
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).fill = FILL_TOTAL
            ws.cell(row=row, column=col).font = Font(bold=True)

    ws.cell(row=13, column=1, value="COMPARAÇÃO").font = F_SEC
    comp = [
        (14, "Menor custo entre os três (R$)", "=MIN($B$5:$D$5)", MOEDA, ""),
        (15, "Diferença AÉREO − RODOVIÁRIO (R$)", "=$B$5-$C$5", MOEDA, "Positivo = o aéreo custa isso a mais."),
        (16, "Diferença AÉREO − RODOVIÁRIO (%)", "=IFERROR(($B$5-$C$5)/$C$5,0)", PCT, "Sobre o rodoviário."),
        (17, "Diferença AÉREO − FRETAMENTO (R$)", "=$B$5-$D$5", MOEDA, ""),
        (18, "Diferença AÉREO − FRETAMENTO (%)", "=IFERROR(($B$5-$D$5)/$D$5,0)", PCT, "Sobre o fretamento."),
        (19, "Diferença RODOVIÁRIO − FRETAMENTO (R$)", "=$C$5-$D$5", MOEDA, ""),
        (20, "Diferença RODOVIÁRIO − FRETAMENTO (%)", "=IFERROR(($C$5-$D$5)/$D$5,0)", PCT, ""),
    ]
    for row, item, f, fmt, obs in comp:
        ws.cell(row=row, column=1, value=item).border = BORDA
        c = ws.cell(row=row, column=2, value=f); c.border = BORDA; c.number_format = fmt; c.font = Font(bold=True)
        o = ws.cell(row=row, column=5, value=obs); o.font = Font(size=9, color="666666")

    ws.cell(row=22, column=1, value="VEREDITO").font = F_SEC
    ver = [
        (23, "Avião x ônibus de linha",
         '=IF($B$5<$C$5,"AÉREO mais barato","RODOVIÁRIO mais barato")'),
        (24, "O mais barato dos três",
         '=IF($B$5=MIN($B$5:$D$5),"AÉREO é o mais barato",IF($C$5=MIN($B$5:$D$5),"RODOVIÁRIO é o mais barato","FRETAMENTO é o mais barato"))'),
        (25, "Alguma opção cabe no orçamento?",
         f'=IF(MIN($B$5:$D$5)<={P}$B$24,"SIM — a mais barata cabe, com folga de "&TEXT({P}$B$24-MIN($B$5:$D$5),"R$ #,##0.00"),'
         f'"NÃO — nem a mais barata cabe. Falta "&TEXT(MIN($B$5:$D$5)-{P}$B$24,"R$ #,##0.00"))'),
        (26, "Ilusão do ônibus barato (o que a diária + hotel escondiam)",
         '="Sem diária e hotel o ônibus daria "&TEXT(RODOVIARIO!$E$22,"R$ #,##0.00")&" — ou seja, '
         'você esconderia "&TEXT(RODOVIARIO!$E$23,"R$ #,##0.00")&" de si mesmo."'),
        (27, "Preço oficial de bagagem",
         '=AEREO!$A$23'),
    ]
    for row, item, f in ver:
        ws.cell(row=row, column=1, value=item).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=f)
        c.fill = FILL_TOTAL
        c.font = Font(bold=True, size=11, color=AZUL)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 30

    ws.merge_cells(start_row=29, start_column=1, end_row=29, end_column=5)
    c = ws.cell(row=29, column=1, value="Custo não é a única variável: o aéreo devolve horas de equipe e reduz risco de atraso na passagem de som; "
                                        "o fretado carrega volume que o aéreo recusa (a Azul, por exemplo, trata amplificador e caixa acústica como "
                                        "artigo perigoso — só até 95 cm somados e 1 por passageiro; acima disso vira Azul Cargo). "
                                        "Decida com o número na mão, não no lugar dele.")
    c.font = Font(italic=True, size=9, color="666666"); c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[29].height = 44
    larguras(ws, {"A": 44, "B": 20, "C": 20, "D": 20, "E": 44})
    ws.freeze_panes = "A5"

    aba_tabelas_oficiais(wb)

    path = os.path.join(OUT_DIR, "SMU_Comparador_Aviao_x_Onibus.xlsx")
    wb.save(path)
    return path


# ================================================================
# PLANILHA 2 - BAGAGEM DA COMITIVA
# ================================================================
LIN_COM_INI = 6
LIN_COM_FIM = 45   # folga generosa: 40 linhas de comitiva


def planilha_bagagem():
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- PREMISSAS ----------------
    ws = wb.create_sheet("PREMISSAS")
    titulo(ws, "PREMISSAS — preencha SÓ as células AMARELAS", 5)
    nota(ws, 2, "A companhia e a antecedência moram aqui, e SÓ aqui. Toda a aba COMITIVA lê estas células "
                "com referência absoluta ($B$5). Trocou a cia? Troca em B5 e a comitiva inteira recalcula o custo. "
                "AMARELO = você escolhe | CINZA = fórmula, não digite por cima.", 5)
    ws.row_dimensions[2].height = 40
    header(ws, 4, ["Premissa", "Valor", "Unidade", "Origem do dado", "Por que isso importa"])

    prem = [
        (5, "Companhia aérea", "GOL", "lista", "ESCOLHA (lista suspensa)",
         "LATAM não publica preço de despacho — é dinâmico. Escolha LATAM e veja a planilha avisar em vez de inventar número.", None, "in"),
        (6, "Antecedência da compra do despacho", "48h ou mais", "lista", "ESCOLHA (lista suspensa)",
         "GOL: 1ª peça R$ 130 antecipado × R$ 180 no balcão do aeroporto. Comprar bagagem em cima da hora é dinheiro jogado fora.", None, "in"),
        (7, "Nº de trechos (ida e volta = 2)", 2, "trechos", "PREMISSA DO ALUNO",
         "Bagagem se paga POR TRECHO. Orçar só a ida é o erro mais comum de comitiva.", '0', "in"),
    ]
    for row, label, valor, unid, origem, porque, fmt, tipo in prem:
        ws.cell(row=row, column=1, value=label).border = BORDA
        c = ws.cell(row=row, column=2, value=valor)
        c.fill = FILL_INPUT; c.border = BORDA; c.font = Font(bold=True)
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=origem); o.border = BORDA
        o.alignment = Alignment(wrap_text=True, vertical="top"); o.font = Font(size=9, color="B71C1C")
        p = ws.cell(row=row, column=5, value=porque); p.border = BORDA
        p.alignment = Alignment(wrap_text=True, vertical="top"); p.font = Font(size=9, color="666666")
        ws.row_dimensions[row].height = 34
        c.comment = Comment(f"{origem}\n\n{porque}", "SMU - Pleno Produtor")

    ws.cell(row=9, column=1, value="CALCULADO (não digite por cima)").font = F_SEC
    calc = [
        (10, "Nº de pessoas na comitiva", f"=COUNTA(COMITIVA!$A${LIN_COM_INI}:$A${LIN_COM_FIM})", "pessoas",
         "FÓRMULA — conta os nomes da aba COMITIVA. Uma célula, uma verdade: não se digita gente aqui, se digita gente lá.", '0'),
        (11, "Chave de consulta da tabela oficial", '=$B$5&"|"&$B$6', "texto",
         "FÓRMULA — é a chave que o ÍNDICE+CORRESP usa em TABELAS_OFICIAIS", None),
        (12, "Preço da 1ª peça despachada (resolvido)", preco_peca("$B$11", 1), "R$/peça/trecho",
         "TABELAS_OFICIAIS — dado oficial, snapshot 15/07/2026", MOEDA),
        (13, "Preço da 2ª peça despachada (resolvido)", preco_peca("$B$11", 2), "R$/peça/trecho",
         "TABELAS_OFICIAIS — dado oficial, snapshot 15/07/2026", MOEDA),
        (14, "Preço da 3ª a 5ª peça (resolvido)", preco_peca("$B$11", 3), "R$/peça/trecho",
         "TABELAS_OFICIAIS — dado oficial, snapshot 15/07/2026", MOEDA),
        (15, "Preço do excesso 32–45 kg (resolvido)", preco_excesso("$B$5"), "R$/volume/trecho",
         "TABELAS_OFICIAIS — só a GOL publica (R$ 350)", MOEDA),
        (16, "Preço da bagagem especial/diferenciada (resolvido)", preco_especial("$B$5"), "R$/volume/trecho",
         "TABELAS_OFICIAIS — GOL R$ 195 (até 292 cm) | Azul R$ 250 (acima de 158 cm) | LATAM não publicado", MOEDA),
    ]
    for row, label, f, unid, origem, fmt in calc:
        ws.cell(row=row, column=1, value=label).border = BORDA
        c = ws.cell(row=row, column=2, value=f)
        c.fill = FILL_CALC; c.border = BORDA; c.font = Font(bold=True)
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=origem); o.border = BORDA
        o.alignment = Alignment(wrap_text=True, vertical="top"); o.font = Font(size=9, color="1B5E20")
        ws.row_dimensions[row].height = 30
        c.comment = Comment(origem, "SMU - Pleno Produtor")

    ws.cell(row=18, column=1, value="LIMITES QUE A ABA COMITIVA VIGIA (dado oficial — snapshot 15/07/2026)").font = F_SEC
    lim = [
        (19, "Peso da franquia por volume despachado", 23, "kg", "GOL / Azul / LATAM — 23 kg por volume"),
        (20, "Peso acima do qual a cia RECUSA o volume", 45, "kg", "GOL 45 kg BR / Azul 45 kg BR / LATAM não transporta acima de 45 kg. "
                                                                   "Acima disso: contrato de CARGA (Res. ANAC 400/2016, art. 15)"),
        (21, "Peso a partir do qual há excesso cobrável", 32, "kg", "Faixa publicada: 32→45 kg. A faixa 23→32 kg da GOL NÃO entra: "
                                                                    "duas páginas oficiais divergem (R$ 170 × R$ 275)"),
        (22, "Franquia mínima de bagagem de mão", 10, "kg", "Res. ANAC 400/2016, art. 14 — mínimo 10 kg. "
                                                            "A norma NÃO fixa dimensão: 55×35×25 cm é contrato da cia, não 'padrão ANAC'"),
        (23, "Máximo de peças despachadas por passageiro/trecho", 5, "peças", "Azul — máx. 5 peças de 23 kg por passageiro/trecho"),
    ]
    for row, label, valor, unid, origem in lim:
        ws.cell(row=row, column=1, value=label).border = BORDA
        c = ws.cell(row=row, column=2, value=valor)
        c.fill = FILL_OFICIAL; c.border = BORDA; c.font = Font(bold=True); c.number_format = '0'
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=origem); o.border = BORDA
        o.alignment = Alignment(wrap_text=True, vertical="top"); o.font = Font(size=9, color="1B5E20")
        ws.row_dimensions[row].height = 32
        c.comment = Comment("DADO OFICIAL — " + origem, "SMU - Pleno Produtor")

    dv_cia = DataValidation(type="list", formula1='"GOL,Azul,LATAM"', allow_blank=False)
    dv_cia.error = "Escolha GOL, Azul ou LATAM."
    dv_cia.errorTitle = "Companhia inválida"
    ws.add_data_validation(dv_cia)
    dv_cia.add(ws["B5"])

    dv_ant = DataValidation(type="list", formula1='"48h ou mais,menos de 48h,no aeroporto"', allow_blank=False)
    dv_ant.error = "Escolha: 48h ou mais / menos de 48h / no aeroporto."
    dv_ant.errorTitle = "Antecedência inválida"
    ws.add_data_validation(dv_ant)
    dv_ant.add(ws["B6"])

    larguras(ws, {"A": 48, "B": 16, "C": 20, "D": 46, "E": 46})
    ws.freeze_panes = "A5"

    # ---------------- COMITIVA ----------------
    ws = wb.create_sheet("COMITIVA")
    titulo(ws, "COMITIVA — quem leva o quê, quanto pesa, quanto custa", 16)
    nota(ws, 2, "Preencha as colunas AMARELAS (A a I). Da J em diante é tudo fórmula — não digite por cima. "
                "Vermelho claro = volume acima de 23 kg (você vai pagar excesso). Vermelho forte = acima de 45 kg: "
                "a companhia RECUSA, isso deixa de ser bagagem e vira contrato de carga (Res. ANAC 400/2016, art. 15). "
                "O custo sai por ÍNDICE+CORRESP na aba TABELAS_OFICIAIS, com a cia e a antecedência que estão em PREMISSAS.", 16)
    ws.row_dimensions[2].height = 52
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=16)
    c = ws.cell(row=3, column=1, value="Bagagem de mão: franquia mínima de 10 kg (Res. ANAC 400/2016, art. 14). "
                                       "Dimensão NÃO é da norma — é do contrato da cia (GOL/LATAM 55×35×25 até 12 kg; Azul 55×35×25 / 115 cm até 10 kg). "
                                       "Case rígido (bumbo, contrabaixo, tuba, violoncelo) = bagagem ESPECIAL/DIFERENCIADA: use a coluna própria. "
                                       "Amplificador e caixa acústica na Azul são ARTIGO PERIGOSO: só até 95 cm somados e 1 por passageiro — acima vai de Azul Cargo.")
    c.font = Font(size=9, color="B71C1C"); c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 46

    cols = ["Nome", "Função", "Mala de mão (kg)", "Volumes despachados",
            "Peso vol. 1 (kg)", "Peso vol. 2 (kg)", "Peso vol. 3 (kg)", "Peso vol. 4 (kg)", "Peso vol. 5 (kg)",
            "Volumes especiais (case rígido)", "Peso total despachado (kg)", "Volume mais pesado (kg)",
            "Alerta mala de mão", "Alerta excesso (>23 kg)", "Alerta recusa (>45 kg)", "Custo do despacho (R$)"]
    header(ws, 5, cols)

    exemplo = [
        ("Ana Ribeiro", "Produtora", 9, 1, 20, None, None, None, None, 0),
        ("Carlos Menezes", "Técnico de som", 10, 2, 23, 28, None, None, None, 0),
        ("Diego Alves", "Baterista", 8, 1, 23, None, None, None, None, 1),
        ("Fernanda Luz", "Iluminadora", 12, 1, 23, None, None, None, None, 0),
        ("Gustavo Pinto", "Roadie", 10, 3, 23, 23, 34, None, None, 0),
        ("Helena Braga", "Backing vocal", 7, 1, 18, None, None, None, None, 0),
        ("Ivan Cordeiro", "Contrabaixista", 9, 1, 23, None, None, None, None, 1),
        ("Júlia Prado", "Assistente de palco", 8, 1, 21, None, None, None, None, 0),
    ]

    PR = "PREMISSAS!"
    for i in range(LIN_COM_INI, LIN_COM_FIM + 1):
        idx = i - LIN_COM_INI
        dados = exemplo[idx] if idx < len(exemplo) else (None,) * 10
        for j in range(10):
            c = ws.cell(row=i, column=1 + j, value=dados[j])
            c.border = BORDA
            c.fill = FILL_INPUT
            if j >= 2:
                c.number_format = '0.#'

        # K: peso total despachado
        c = ws.cell(row=i, column=11, value=f"=IF($A{i}=\"\",\"\",SUM($E{i}:$I{i}))")
        c.border = BORDA; c.number_format = '0.#'; c.fill = FILL_CALC
        # L: volume mais pesado
        c = ws.cell(row=i, column=12, value=f'=IF($A{i}="","",IF(COUNT($E{i}:$I{i})=0,0,MAX($E{i}:$I{i})))')
        c.border = BORDA; c.number_format = '0.#'; c.fill = FILL_CALC
        # M: alerta mala de mao
        c = ws.cell(row=i, column=13,
                    value=f'=IF($A{i}="","",IF($C{i}>{PR}$B$22,"acima de "&{PR}$B$22&" kg — confira o contrato da cia","ok"))')
        c.border = BORDA; c.font = Font(size=9); c.fill = FILL_CALC
        # N: alerta excesso
        c = ws.cell(row=i, column=14,
                    value=f'=IF($A{i}="","",IF($L{i}>{PR}$B$19,"EXCESSO","ok"))')
        c.border = BORDA; c.font = Font(size=9, bold=True); c.fill = FILL_CALC
        # O: alerta recusa
        c = ws.cell(row=i, column=15,
                    value=f'=IF($A{i}="","",IF($L{i}>{PR}$B$20,"RECUSADO — vira carga",""))')
        c.border = BORDA; c.font = Font(size=9, bold=True); c.fill = FILL_CALC
        # P: custo do despacho (INDEX+MATCH inline, ida+volta)
        f = (
            f'=IF($A{i}="","",'
            f'IFERROR({PR}$B$7*('
            f'IF($D{i}>=1,INDEX({RNG_PRECOS},MATCH({PR}$B$11,{RNG_CHAVE},0),1),0)'
            f'+IF($D{i}>=2,INDEX({RNG_PRECOS},MATCH({PR}$B$11,{RNG_CHAVE},0),2),0)'
            f'+IF($D{i}>=3,MAX(0,MIN($D{i},5)-2)*INDEX({RNG_PRECOS},MATCH({PR}$B$11,{RNG_CHAVE},0),3),0)'
            f'+IF(COUNTIFS($E{i}:$I{i},">"&{PR}$B$21,$E{i}:$I{i},"<="&{PR}$B$20)>0,'
            f'COUNTIFS($E{i}:$I{i},">"&{PR}$B$21,$E{i}:$I{i},"<="&{PR}$B$20)*INDEX({RNG_EXC_VAL},MATCH({PR}$B$5,{RNG_EXC_CIA},0)),0)'
            f'+IF($J{i}>0,$J{i}*INDEX({RNG_ESP_VAL},MATCH({PR}$B$5,{RNG_ESP_CIA},0)),0)'
            f'),"n/d — cotar com a cia"))'
        )
        c = ws.cell(row=i, column=16, value=f)
        c.border = BORDA; c.number_format = MOEDA; c.fill = FILL_TOTAL; c.font = Font(bold=True)

    ws.cell(row=LIN_COM_INI, column=16).comment = Comment(
        "Custo = nº de trechos × ( 1ª peça + 2ª peça + (nº de peças além da 2ª, até 5) × preço da 3ª–5ª "
        "+ nº de volumes na faixa 32–45 kg × preço do excesso + volumes especiais × preço da bagagem diferenciada ).\n\n"
        "Todos os preços saem de TABELAS_OFICIAIS por ÍNDICE+CORRESP, usando a chave montada em PREMISSAS!B11.\n"
        "Se a cia escolhida não publica o preço (LATAM), a fórmula devolve 'n/d — cotar com a cia' em vez de inventar número.",
        "SMU - Pleno Produtor")

    # formatacao condicional nos pesos (E:I) e no volume mais pesado (L)
    faixa_peso = f"E{LIN_COM_INI}:I{LIN_COM_FIM}"
    ws.conditional_formatting.add(faixa_peso, CellIsRule(
        operator="greaterThan", formula=[str(45)],
        fill=PatternFill("solid", fgColor="C62828"), font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add(faixa_peso, CellIsRule(
        operator="between", formula=["23.0000001", "45"],
        fill=PatternFill("solid", fgColor="FFCDD2"), font=Font(color="B71C1C", bold=True)))
    faixa_max = f"L{LIN_COM_INI}:L{LIN_COM_FIM}"
    ws.conditional_formatting.add(faixa_max, CellIsRule(
        operator="greaterThan", formula=["45"],
        fill=PatternFill("solid", fgColor="C62828"), font=Font(color="FFFFFF", bold=True)))
    ws.conditional_formatting.add(faixa_max, CellIsRule(
        operator="between", formula=["23.0000001", "45"],
        fill=PatternFill("solid", fgColor="FFCDD2"), font=Font(color="B71C1C", bold=True)))
    # alertas de texto
    ws.conditional_formatting.add(f"N{LIN_COM_INI}:O{LIN_COM_FIM}", CellIsRule(
        operator="equal", formula=['"EXCESSO"'],
        fill=PatternFill("solid", fgColor="FFCDD2"), font=Font(color="B71C1C", bold=True)))
    ws.conditional_formatting.add(f"N{LIN_COM_INI}:O{LIN_COM_FIM}", CellIsRule(
        operator="equal", formula=['"RECUSADO — vira carga"'],
        fill=PatternFill("solid", fgColor="C62828"), font=Font(color="FFFFFF", bold=True)))

    dv_vol = DataValidation(type="whole", operator="between", formula1="0", formula2="5", allow_blank=True)
    dv_vol.error = "Azul: máximo 5 peças de 23 kg por passageiro/trecho (snapshot 15/07/2026)."
    dv_vol.errorTitle = "Volumes despachados"
    ws.add_data_validation(dv_vol)
    dv_vol.add(f"D{LIN_COM_INI}:D{LIN_COM_FIM}")

    larguras(ws, {"A": 22, "B": 20, "C": 14, "D": 13, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12,
                  "J": 14, "K": 14, "L": 13, "M": 30, "N": 14, "O": 22, "P": 18})
    ws.freeze_panes = "C6"

    # ---------------- RESUMO ----------------
    ws = wb.create_sheet("RESUMO")
    titulo(ws, "RESUMO DA BAGAGEM DA COMITIVA", 4)
    nota(ws, 2, "Tudo aqui é fórmula sobre a aba COMITIVA, com folga até a linha 45 — "
                "acrescente gente lá e este resumo acompanha sozinho.", 4)
    header(ws, 4, ["Indicador", "Valor", "Unidade", "Observação"])

    C = "COMITIVA!"
    a, b = LIN_COM_INI, LIN_COM_FIM
    res = [
        (5, "Pessoas na comitiva", f"=COUNTA({C}$A${a}:$A${b})", "pessoas", "Conta os nomes preenchidos. É a mesma célula de PREMISSAS!B10.", '0'),
        (6, "Volumes despachados (por trecho)", f"=SUM({C}$D${a}:$D${b})", "volumes", "", '0'),
        (7, "Volumes despachados (ida + volta)", f"=SUM({C}$D${a}:$D${b})*PREMISSAS!$B$7", "volumes", "Bagagem se paga POR TRECHO.", '0'),
        (8, "Volumes especiais / case rígido (por trecho)", f"=SUM({C}$J${a}:$J${b})", "volumes", "GOL R$ 195 (até 292 cm) | Azul R$ 250 (acima de 158 cm) | LATAM não publica.", '0'),
        (9, "Peso total despachado (por trecho)", f"=SUM({C}$K${a}:$K${b})", "kg", "", '#,##0.0'),
        (10, "Peso total de mala de mão", f"=SUM({C}$C${a}:$C${b})", "kg", "Franquia mínima de 10 kg por pessoa (Res. ANAC 400/2016, art. 14).", '#,##0.0'),
        (11, "Peso médio por volume despachado", f"=IFERROR(SUM({C}$K${a}:$K${b})/SUM({C}$D${a}:$D${b}),0)", "kg", "Acima de 23 kg você já paga excesso em algum volume.", '0.0'),
        (12, "Volume mais pesado da comitiva", f"=IFERROR(MAX({C}$E${a}:$I${b}),0)", "kg", "Se passar de 45 kg, a cia recusa: vira contrato de carga.", '0.0'),
    ]
    for row, item, f, unid, obs, fmt in res:
        ws.cell(row=row, column=1, value=item).border = BORDA
        ws.cell(row=row, column=1).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=f); c.border = BORDA; c.number_format = fmt; c.font = Font(bold=True)
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=obs); o.border = BORDA
        o.font = Font(size=9, color="666666"); o.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 26

    ws.cell(row=14, column=1, value="CUSTO").font = F_SEC
    custo = [
        (15, "CUSTO TOTAL DA BAGAGEM (ida + volta)", f"=SUM({C}$P${a}:$P${b})", "R$",
         "SOMA(COMITIVA!P6:P45) — com folga até a linha 45 de propósito. A SOMA ignora texto em silêncio: "
         "se a cia não publicar preço, o alerta abaixo é o que te salva.", MOEDA),
        (16, "Custo médio de bagagem por pessoa", f"=IFERROR(SUM({C}$P${a}:$P${b})/COUNTA({C}$A${a}:$A${b}),0)", "R$", "", MOEDA),
        (17, "Custo médio por volume despachado", f"=IFERROR(SUM({C}$P${a}:$P${b})/(SUM({C}$D${a}:$D${b})*PREMISSAS!$B$7),0)", "R$",
         "Compare com o preço da 1ª peça: se estiver bem acima, você está pagando excesso e não percebeu.", MOEDA),
        (18, "Custo se TODO mundo despachasse 1 peça só, sem excesso e sem especial",
         f"=IFERROR(COUNTA({C}$A${a}:$A${b})*PREMISSAS!$B$12*PREMISSAS!$B$7,0)", "R$",
         "É o piso teórico da comitiva. A diferença para a linha 15 é o que a bagagem extra, o excesso e o case rígido custam de verdade.", MOEDA),
        (19, "Quanto a bagagem extra está custando além do piso", f"=IFERROR(SUM({C}$P${a}:$P${b})-$B$18,0)", "R$",
         "Se este número te assusta, é porque ele deveria ter aparecido no orçamento — e não na fatura.", MOEDA),
    ]
    for row, item, f, unid, obs, fmt in custo:
        ws.cell(row=row, column=1, value=item).border = BORDA
        ws.cell(row=row, column=1).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=f); c.border = BORDA; c.number_format = fmt
        c.font = Font(bold=True, size=12 if row == 15 else 10)
        if row == 15:
            c.fill = FILL_TOTAL
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=obs); o.border = BORDA
        o.font = Font(size=9, color="666666"); o.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28

    ws.cell(row=20, column=1, value="ALERTAS").font = F_SEC
    al = [
        (21, "Nº de volumes com EXCESSO (acima de 23 kg)", f'=COUNTIF({C}$N${a}:$N${b},"EXCESSO")', "pessoas",
         "Conta PESSOAS cujo volume mais pesado passa de 23 kg."),
        (22, "Nº de volumes RECUSADOS (acima de 45 kg)", f'=COUNTIF({C}$O${a}:$O${b},"RECUSADO — vira carga")', "pessoas",
         "Acima de 45 kg não é bagagem: é carga. Res. ANAC 400/2016, art. 15."),
        (23, "Nº de pessoas com mala de mão acima de 10 kg", f'=COUNTIF({C}$C${a}:$C${b},">"&PREMISSAS!$B$22)', "pessoas",
         "10 kg é a franquia MÍNIMA da norma. Acima disso depende do contrato da cia — confira antes."),
        (24, "Total de alertas", "=$B$21+$B$22+$B$23", "alertas", "Zero é a meta. Não é decoração."),
    ]
    for row, item, f, unid, obs in al:
        ws.cell(row=row, column=1, value=item).border = BORDA
        c = ws.cell(row=row, column=2, value=f); c.border = BORDA; c.number_format = '0'; c.font = Font(bold=True)
        ws.cell(row=row, column=3, value=unid).border = BORDA
        o = ws.cell(row=row, column=4, value=obs); o.border = BORDA
        o.font = Font(size=9, color="666666"); o.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 26
    ws.conditional_formatting.add("B21:B24", CellIsRule(
        operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFCDD2"), font=Font(color="B71C1C", bold=True)))
    ws.conditional_formatting.add("B22", CellIsRule(
        operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C62828"), font=Font(color="FFFFFF", bold=True)))

    diag = [
        (26, f'=IF(COUNTIF({C}$P${a}:$P${b},"n/d — cotar com a cia")>0,'
             f'"⚠ A companhia escolhida NÃO publica preço de despacho. A SOMA da linha 15 está ignorando essas linhas em SILÊNCIO — '
             f'o total está mais baixo que a realidade. Cote com a cia antes de fechar.","ok — todos os itens têm preço oficial publicado")'),
        (27, f'=IF($B$22>0,"⛔ Há volume acima de 45 kg. Isso não embarca como bagagem: a companhia recusa ou submete a contrato de transporte de CARGA '
             f'(Res. ANAC 400/2016, art. 15). Replaneje ou mande de carga.","")'),
        (28, f'=IF($B$21>0,"⚠ Há volume na faixa acima de 23 kg. Entre 23 e 32 kg a GOL tem DUAS páginas oficiais divergentes (R$ 170 × R$ 275) — '
             f'esta planilha não cobra essa faixa de propósito, então o custo real pode ser MAIOR. De 32 a 45 kg a GOL publica R$ 350/volume.","")'),
        (29, f'=IF(SUM({C}$J${a}:$J${b})>0,"⚠ Há bagagem especial/diferenciada (case rígido). Lembre: na GOL, cliente Smiles NÃO tem gratuidade, '
             f'tarifa com despacho grátis NÃO cobre a diferenciada, e a GOL não se responsabiliza por danos a bagagem diferenciada.","")'),
        (30, f'=IF(PREMISSAS!$B$5="Azul","⚠ Azul: alto-falante, amplificador e caixa acústica são ARTIGO PERIGOSO — só até 95 cm somados e 1 por passageiro. '
             f'Acima disso vai de Azul Cargo, não de bagagem. É isto que quebra o plano de levar backline na mala.","")'),
        (31, f'=IF(PREMISSAS!$B$7<2,"⚠ Você está orçando bagagem em 1 trecho só. Despacho se paga POR TRECHO: ida e volta = 2.","")'),
        (32, '="Lembrete: o PL 5041/2025 (12 kg de mão + 23 kg despachado grátis) NÃO é lei. Foi aprovado na Câmara em 28/10/2025 e está parado '
             'no Senado desde 05/11/2025. Blog que diz que já vale está errado — e quem orça por ele perde dinheiro."'),
    ]
    for row, f in diag:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f)
        c.fill = FILL_ALERTA
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.font = Font(size=9, bold=True, color="B71C1C")
        ws.row_dimensions[row].height = 30

    larguras(ws, {"A": 46, "B": 18, "C": 12, "D": 60})
    ws.freeze_panes = "A5"

    aba_tabelas_oficiais(wb)

    path = os.path.join(OUT_DIR, "SMU_Bagagem_da_Comitiva.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    p1 = planilha_comparador()
    print("OK:", p1)
    p2 = planilha_bagagem()
    print("OK:", p2)
